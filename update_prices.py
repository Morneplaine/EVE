"""
Update Prices sheet in existing Excel file with latest Fuzzwork Market data
This script reads the existing Excel file, extracts all typeIDs, fetches new prices, and updates the Prices sheet

Usage:
    python update_prices.py [excel_file.xlsx]
    
If no file is specified, it will look for EVE_Manufacturing_Database.xlsx in the current directory.
"""

import sys
import pandas as pd
from pathlib import Path
from eve_manufacturing_database import get_fuzzwork_market_prices, JITA_SYSTEM_ID
import logging
from openpyxl import load_workbook

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def update_prices_in_excel(excel_file="EVE_Manufacturing_Database.xlsx"):
    """
    Update the Prices sheet in an existing Excel file with latest market data.
    
    Args:
        excel_file (str): Path to the Excel file to update
    """
    if not Path(excel_file).exists():
        logger.error(f"Excel file not found: {excel_file}")
        logger.error("Please run the main script first to create the database.")
        return False
    
    logger.info(f"Reading existing Excel file: {excel_file}")
    
    try:
        # Read the Prices sheet to get all typeIDs
        prices_df = pd.read_excel(excel_file, sheet_name='Prices')
        type_ids = prices_df['typeID'].tolist()
        logger.info(f"Found {len(type_ids)} items in Prices sheet")
        
        # Fetch latest prices from Fuzzwork Market API
        logger.info("Fetching latest prices from Fuzzwork Market API for Jita...")
        logger.info("This may take a few minutes...")
        fuzzwork_prices = get_fuzzwork_market_prices(type_ids, station_id=JITA_SYSTEM_ID)
        
        # Update the prices DataFrame
        logger.info("Updating prices in DataFrame...")
        updated_count = 0
        for idx, row in prices_df.iterrows():
            type_id = int(row['typeID'])
            if type_id in fuzzwork_prices:
                fw_price = fuzzwork_prices[type_id]
                prices_df.at[idx, 'Buy Max'] = fw_price.get('buy_max', 0)
                prices_df.at[idx, 'Buy Volume'] = fw_price.get('buy_volume', 0)
                prices_df.at[idx, 'Sell Min'] = fw_price.get('sell_min', 0)
                prices_df.at[idx, 'Sell Avg'] = fw_price.get('sell_min', 0)  # Use sell_min as avg
                prices_df.at[idx, 'Sell Median'] = fw_price.get('sell_min', 0)
                prices_df.at[idx, 'Sell Volume'] = fw_price.get('sell_volume', 0)
                if fw_price.get('buy_max', 0) > 0 or fw_price.get('sell_min', 0) > 0:
                    updated_count += 1
        
        # Load the workbook and update only the Prices sheet
        logger.info(f"Writing updated prices to {excel_file}...")
        workbook = load_workbook(excel_file)
        
        # Remove old Prices sheet if it exists
        if 'Prices' in workbook.sheetnames:
            del workbook['Prices']
        
        # Create new Prices sheet
        from openpyxl.utils.dataframe import dataframe_to_rows
        ws = workbook.create_sheet('Prices')
        
        # Write headers
        headers = list(prices_df.columns)
        ws.append(headers)
        
        # Write data
        for r in dataframe_to_rows(prices_df, index=False, header=False):
            ws.append(r)
        
        # Save the workbook
        workbook.save(excel_file)
        workbook.close()
        
        logger.info("=" * 60)
        logger.info("SUCCESS! Prices updated successfully")
        logger.info(f"Updated prices for {updated_count} items")
        logger.info("The VLOOKUP formulas in Manufacturing and Reprocessing sheets will automatically update")
        logger.info("=" * 60)
        return True
        
    except Exception as e:
        logger.error(f"Error updating prices: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def main():
    excel_file = sys.argv[1] if len(sys.argv) > 1 else "EVE_Manufacturing_Database.xlsx"
    logger.info("=" * 60)
    logger.info("EVE Manufacturing Database - Price Updater")
    logger.info("=" * 60)
    logger.info("This script updates ONLY the Prices sheet in your Excel file")
    logger.info("All other data (manufacturing, reprocessing, volumes) remains unchanged")
    logger.info("=" * 60)
    success = update_prices_in_excel(excel_file)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()

